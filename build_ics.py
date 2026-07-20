"""
Downloads the live MICA PGP-2 Term-3 schedule workbook from SharePoint,
reads the "Schedule Term-3 phase-1" sheet, filters to a chosen list of
course codes, and writes a fresh ICS calendar file.

Run manually:
    python build_ics.py

Configuration is via environment variables (set as GitHub Actions secrets/vars):
    SHAREPOINT_URL   - the SharePoint share link to the .xlsx file
    SHEET_NAME       - exact sheet/tab name containing the schedule
                        (default: "Schedule Term-3 phase-1")
"""

import os
import io
import sys
from datetime import datetime

import requests
import openpyxl
from icalendar import Calendar, Event

# ---- CONFIG: the course codes to keep -------------------------------------
TARGET_CODES = [
    "F1-BDB",
    "F1-OBM",
    "F1-MAPC",
    "F1-CAA",
    "F1-TBO",  # covers F1-TBO:CP
    "F4-MTAI:TNOSM",
    "F4-PMG:EFF",
    "F4-BSCIAIW",
    "F4-CI:CRMDGP",
    "F4-MTAIP:EPF",
]

SESSION_COL_START = 3  # column D (0-indexed 3) = Session 1
NUM_SESSIONS = 8

# Fixed 24-hour session times (the sheet has no AM/PM markers and is
# ambiguous to auto-parse around midday, so these are hardcoded from the
# header row and assumed constant for the whole term).
FIXED_SESSION_TIMES = [
    (9, 15, 10, 30),   # Session 1
    (10, 45, 12, 0),   # Session 2
    (12, 15, 13, 30),  # Session 3
    (14, 15, 15, 30),  # Session 4
    (15, 45, 17, 0),   # Session 5
    (17, 30, 18, 45),  # Session 6
    (19, 0, 20, 15),   # Session 7
    (21, 0, 22, 15),   # Session 8
]

OUTPUT_PATH = "docs/schedule.ics"


def to_download_url(share_url):
    """Convert a SharePoint 'view' share link into a direct-download link."""
    if "download=1" in share_url:
        return share_url
    sep = "&" if "?" in share_url else "?"
    return f"{share_url}{sep}download=1"


def fetch_workbook(share_url):
    url = to_download_url(share_url)
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)


def matches_target(cell_text):
    for code in TARGET_CODES:
        if code in cell_text:
            return code
    return None


def build_events(ws):
    rows = list(ws.iter_rows(values_only=True))
    rows = [["" if c is None else str(c) for c in row] for row in rows]

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0].strip() == "Date":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row starting with 'Date'")

    events = []
    current_date = None
    for row in rows[header_idx + 1:]:
        if not any(cell.strip() for cell in row):
            continue

        date_str = row[0].strip()
        classroom = row[2].strip() if len(row) > 2 else ""

        if date_str:
            parsed = None
            for fmt in ("%d-%b-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed:
                current_date = parsed

        if current_date is None:
            continue

        for s in range(NUM_SESSIONS):
            col = SESSION_COL_START + s
            if col >= len(row):
                continue
            cell = row[col].strip()
            if not cell:
                continue
            code = matches_target(cell)
            if not code:
                continue
            h1, m1, h2, m2 = FIXED_SESSION_TIMES[s]
            start_dt = datetime.combine(current_date, datetime.min.time()).replace(hour=h1, minute=m1)
            end_dt = datetime.combine(current_date, datetime.min.time()).replace(hour=h2, minute=m2)
            events.append(
                {
                    "start": start_dt,
                    "end": end_dt,
                    "title": cell,
                    "location": classroom,
                    "uid_key": f"{current_date.isoformat()}-S{s + 1}-{code}",
                }
            )

    return events


def write_ics(events, out_path):
    cal = Calendar()
    cal.add("prodid", "-//MICA Schedule Sync//mxm.dk//")
    cal.add("version", "2.0")
    cal.add("X-WR-CALNAME", "MICA Term-3 Schedule (Filtered)")
    cal.add("REFRESH-INTERVAL;VALUE=DURATION", "PT15M")

    for ev in events:
        e = Event()
        e.add("summary", ev["title"])
        e.add("dtstart", ev["start"])
        e.add("dtend", ev["end"])
        if ev["location"]:
            e.add("location", ev["location"])
        e.add("uid", ev["uid_key"] + "@mica-schedule-sync")
        cal.add_component(e)

    dirpath = os.path.dirname(out_path)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)
    with open(out_path, "wb") as f:
        f.write(cal.to_ical())


if __name__ == "__main__":
    share_url = os.environ.get("SHAREPOINT_URL")
    sheet_name = os.environ.get("SHEET_NAME", "Schedule Term-3 phase-1")

    if not share_url:
        print("ERROR: set the SHAREPOINT_URL environment variable.", file=sys.stderr)
        sys.exit(1)

    wb = fetch_workbook(share_url)
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    events = build_events(ws)
    write_ics(events, OUTPUT_PATH)
    print(f"Wrote {len(events)} events to {OUTPUT_PATH}")
